

from __future__ import annotations

import logging
import sys

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from scraper.config import ScraperConfig
from scraper.exceptions import WebFormsError
from scraper.forms import parse_html
from scraper.normalize import parse_date
from scraper.parse_pages import has_next_page, parse_student_detail, parse_student_grid
from scraper.webforms_client import WebFormsClient

LOGIN_PATH = "/Login.aspx"
LIST_PATH = "/Default.aspx"
DETAIL_PATH = "/Details.aspx"

HEADER = [
    "Student ID", "Name", "Campus", "Course", "Enrolled", "Status",
    "Email", "Phone", "Date of Birth", "Address", "Emergency Contact",
]


def _all_campus_options(html: str) -> list[tuple[str, str]]:

    soup = parse_html(html)
    select = soup.find("select", id=lambda v: bool(v) and "ddlCampus" in v)
    options: list[tuple[str, str]] = []
    if select:
        for opt in select.find_all("option"):
            value = opt.get("value", "")
            options.append((value if isinstance(value, str) else "", opt.get_text(strip=True)))
    return options


def scrape_all(client: WebFormsClient) -> list[dict[str, object]]:
    results: list[dict[str, object]] = []
    seen_ids: set[str] = set()

    list_html = client.login("demo", "demo123", login_path=LOGIN_PATH)
    list_html = client.get(LIST_PATH)

    campuses = _all_campus_options(list_html)
    logging.info("Found %d campus filter options", len(campuses))

    for campus_value, campus_label in campuses:
        logging.info("Scraping campus filter: %s", campus_label or "(all)")

        if campus_value:
            page_html = client.postback(
                list_html, LIST_PATH, target_hint="ddlCampus"
            )
        else:
            page_html = list_html

        page_number = 1
        while True:
            rows = parse_student_grid(page_html)
            logging.info("  Page %d: %d rows", page_number, len(rows))

            for row in rows:
                if row.student_id in seen_ids:
                    continue  # a student can appear under "All" and a named campus
                seen_ids.add(row.student_id)

                detail_html = client.postback(
                    page_html, LIST_PATH, target_hint=row.detail_target
                )
                detail = parse_student_detail(detail_html)

                results.append(
                    {
                        "student_id": row.student_id,
                        "name": row.name,
                        "campus": row.campus,
                        "course": row.course,
                        "enrolled": parse_date(row.enrolled),
                        "status": row.status,
                        "email": detail.email,
                        "phone": detail.phone,
                        "date_of_birth": parse_date(detail.date_of_birth),
                        "address": detail.address,
                        "emergency_contact": detail.emergency_contact,
                    }
                )

            if not has_next_page(page_html):
                break
            page_html = client.postback(
                page_html, LIST_PATH, target_hint="gvStudents", argument=f"Page${page_number + 1}"
            )
            page_number += 1

    return results


def write_excel(rows: list[dict[str, object]], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    assert isinstance(ws, Worksheet)
    ws.title = "Students"
    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(
            [
                row["student_id"], row["name"], row["campus"], row["course"],
                row["enrolled"], row["status"], row["email"], row["phone"],
                row["date_of_birth"], row["address"], row["emergency_contact"],
            ]
        )

    for column_cells in ws.columns:
        length = max(len(str(c.value)) for c in column_cells if c.value is not None)
        col_idx = column_cells[0].column
        assert col_idx is not None
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(length + 2, 40)

    wb.save(output_path)


def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-8s %(message)s",
        datefmt="%H:%M:%S",
    )
    config = ScraperConfig.from_env()

    try:
        with WebFormsClient(config.base_url) as client:
            client.session.verify = config.verify_ssl
            rows = scrape_all(client)
    except WebFormsError as exc:
        logging.error("Scrape failed: %s", exc)
        return 1

    write_excel(rows, config.output_path)
    logging.info("Wrote %d records to %s", len(rows), config.output_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())